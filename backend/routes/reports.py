"""
Reports & Export Center
========================
Implements all 8 report types with browser view, print, CSV, Excel (.xlsx)
and PDF exports.

Routes
------
GET  /reports/           — hub index
GET  /reports/driver     — driver report        (?format=csv|excel|pdf)
GET  /reports/vehicle    — vehicle report       (?format=)
GET  /reports/contract   — contract report      (?format=)
GET  /reports/payment    — payment report       (?format=, ?period=daily|weekly|monthly|yearly)
GET  /reports/expense    — expenditure report   (?format=)
GET  /reports/capital    — capital report       (?format=)
GET  /reports/audit      — audit report         (?format=)
GET  /reports/archive    — archive report       (?format=, ?tab=drivers|vehicles|contracts)
"""

import csv
import io
from datetime import date, datetime, timedelta

from flask import (
    Blueprint, render_template, request, Response,
    flash, redirect, url_for,
)
from flask_login import login_required
from sqlalchemy import func

from backend.extensions import db
from backend.models.driver import Driver
from backend.models.vehicle import Vehicle
from backend.models.contract import Contract
from backend.models.payment import Payment
from backend.models.expense import Expense, CATEGORIES
from backend.models.audit import AuditLog
from backend.models.capital import CapitalAdjustment
from backend.models.user import User
from backend.utils import parse_date as _pd

reports_bp = Blueprint("reports", __name__)

# ─── Export library flags ─────────────────────────────────────────────────────

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    _XLSX_OK = True
except ImportError:
    _XLSX_OK = False

try:
    from xhtml2pdf import pisa
    _PDF_OK = True
except ImportError:
    _PDF_OK = False


# ─── Shared helpers ───────────────────────────────────────────────────────────

def _fmt(n):
    """Format a number as a plain decimal string (no currency symbol)."""
    try:
        return f"{float(n or 0):.2f}"
    except Exception:
        return "0.00"


def _period_dates(period, date_from_s, date_to_s):
    """Convert a period preset or custom date strings to (date_from, date_to)."""
    today = date.today()
    if period == "daily":
        return today, today
    if period == "weekly":
        monday = today - timedelta(days=today.weekday())
        return monday, today
    if period == "monthly":
        return today.replace(day=1), today
    if period == "yearly":
        return today.replace(month=1, day=1), today
    return _pd(date_from_s), _pd(date_to_s)


def _csv_resp(headers, rows, filename):
    """Return a UTF-8 BOM CSV download response."""
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(headers)
    for row in rows:
        writer.writerow([("" if v is None else str(v)) for v in row])
    return Response(
        "\ufeff" + buf.getvalue(),
        mimetype="text/csv; charset=utf-8-sig",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _xlsx_resp(sheet_title, col_headers, rows, filename, totals=None):
    """Return a professionally formatted .xlsx download response."""
    if not _XLSX_OK:
        flash("Excel export unavailable — openpyxl is not installed.", "warning")
        return redirect(request.referrer or url_for("reports.index"))

    def _col_letter(n):
        s = ""
        while n > 0:
            n, r = divmod(n - 1, 26)
            s = chr(65 + r) + s
        return s

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_title[:31]

    NAV        = "1A2341"
    hdr_fill   = PatternFill("solid", fgColor=NAV)
    hdr_font   = Font(bold=True, color="FFFFFF", size=10)
    title_font = Font(bold=True, size=14)
    sub_font   = Font(italic=True, size=9, color="888888")
    tot_fill   = PatternFill("solid", fgColor="EEF2FF")
    tot_font   = Font(bold=True, size=10)
    center     = Alignment(horizontal="center")
    thin       = Side(style="thin", color="DDDDDD")
    thin_bdr   = Border(bottom=thin, left=thin, right=thin)

    ncols = len(col_headers)
    end   = _col_letter(ncols)

    # Row 1: report title
    ws.merge_cells(f"A1:{end}1")
    c = ws["A1"]
    c.value, c.font, c.alignment = f"THMS — {sheet_title}", title_font, center
    ws.row_dimensions[1].height = 22

    # Row 2: generated timestamp
    ws.merge_cells(f"A2:{end}2")
    c = ws["A2"]
    c.value = f"Generated: {datetime.now().strftime('%d %b %Y at %H:%M')}"
    c.font, c.alignment = sub_font, center

    ws.append([])  # Row 3 spacer

    # Row 4: column headers
    ws.append(col_headers)
    for ci in range(1, ncols + 1):
        cell = ws.cell(row=4, column=ci)
        cell.fill, cell.font, cell.alignment, cell.border = hdr_fill, hdr_font, center, thin_bdr

    # Data rows
    for row_data in rows:
        ws.append(list(row_data))
        rn = ws.max_row
        for ci in range(1, ncols + 1):
            ws.cell(rn, ci).border = thin_bdr

    # Summary section
    if totals:
        ws.append([])
        ws.append(["── Summary"])
        ws.cell(ws.max_row, 1).font = Font(bold=True, size=11, color=NAV)
        for label, val in totals:
            ws.append([label, val])
            rn = ws.max_row
            ws.cell(rn, 1).fill, ws.cell(rn, 1).font = tot_fill, tot_font
            if ncols >= 2:
                ws.cell(rn, 2).fill, ws.cell(rn, 2).font = tot_fill, tot_font

    # Auto-fit column widths
    for col in ws.columns:
        best = 10
        ltr  = col[0].column_letter
        for cell in col:
            try:
                best = max(best, min(len(str(cell.value or "")), 60))
            except Exception:
                pass
        ws.column_dimensions[ltr].width = best + 3

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        buf.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _pdf_resp(tpl, ctx, filename):
    """Render a Jinja2 template to PDF via xhtml2pdf."""
    if not _PDF_OK:
        flash("PDF export unavailable — xhtml2pdf is not installed.", "warning")
        return redirect(request.referrer or url_for("reports.index"))
    html = render_template(tpl, **ctx)
    buf  = io.BytesIO()
    pisa.CreatePDF(io.StringIO(html), dest=buf)
    buf.seek(0)
    return Response(
        buf.getvalue(),
        mimetype="application/pdf",
        headers={"Content-Disposition": f'inline; filename="{filename}"'},
    )


def _today():
    return date.today().isoformat()


def _now():
    return datetime.now()


# ─── Capital summary (mirrors capital module logic) ───────────────────────────

def _capital_summary():
    vehicle_cost = float(
        db.session.query(func.coalesce(func.sum(Vehicle.purchase_price), 0)).scalar() or 0
    )
    extra_exp = float(
        db.session.query(func.coalesce(func.sum(Expense.amount), 0))
        .filter(Expense.is_archived == False).scalar() or 0
    )
    payments_recv = float(
        db.session.query(func.coalesce(func.sum(Payment.amount), 0))
        .filter(Payment.is_archived == False).scalar() or 0
    )
    manual_added = float(
        db.session.query(func.coalesce(func.sum(CapitalAdjustment.amount), 0))
        .filter(CapitalAdjustment.type == "add").scalar() or 0
    )
    manual_withdrawn = float(
        db.session.query(func.coalesce(func.sum(CapitalAdjustment.amount), 0))
        .filter(CapitalAdjustment.type == "withdraw").scalar() or 0
    )
    outstanding = sum(
        c.outstanding_balance
        for c in Contract.query.filter(Contract.status != "archived").all()
    )
    return {
        "vehicle_cost":        vehicle_cost,
        "extra_expenditure":   extra_exp,
        "payments_received":   payments_recv,
        "manual_added":        manual_added,
        "manual_withdrawn":    manual_withdrawn,
        "outstanding_balance": outstanding,
        "net_capital":         vehicle_cost + manual_added - extra_exp + payments_recv - manual_withdrawn,
    }


# ══════════════════════════════════════════════════════════════════════════════
# INDEX
# ══════════════════════════════════════════════════════════════════════════════

@reports_bp.route("/")
@login_required
def index():
    return render_template("reports/index.html")


# ══════════════════════════════════════════════════════════════════════════════
# 1. DRIVER REPORT
# ══════════════════════════════════════════════════════════════════════════════

@reports_bp.route("/driver")
@login_required
def driver():
    search   = request.args.get("q", "").strip()
    status_f = request.args.get("status", "")
    fmt      = request.args.get("format", "")
    page     = request.args.get("page", 1, type=int)

    q = Driver.query
    if search:
        like = f"%{search}%"
        q = q.filter(Driver.full_name.ilike(like) | Driver.phone.ilike(like))
    if status_f:
        q = q.filter(Driver.status == status_f)
    q = q.order_by(Driver.full_name)

    is_export = bool(fmt)
    pag   = None if is_export else q.paginate(page=page, per_page=50, error_out=False)
    items = q.all() if is_export else pag.items

    def _enrich(d):
        c  = d.active_contract or d.contracts.order_by(Contract.id.desc()).first()
        tp = float(db.session.query(func.coalesce(func.sum(Payment.amount), 0))
                   .filter(Payment.driver_id == d.id, Payment.is_archived == False)
                   .scalar() or 0)
        te = float(db.session.query(func.coalesce(func.sum(Expense.amount), 0))
                   .filter(Expense.driver_id == d.id, Expense.is_archived == False)
                   .scalar() or 0)
        return dict(driver=d, contract=c,
                    vehicle=c.vehicle if c else None,
                    total_paid=tp, total_exp=te,
                    outstanding=c.outstanding_balance if c else 0.0)

    rows              = [_enrich(d) for d in items]
    total_paid        = sum(r["total_paid"]  for r in rows)
    total_outstanding = sum(r["outstanding"] for r in rows)

    HDRS = [
        "Driver Name", "Phone", "Address", "Next of Kin", "Guarantor 1",
        "Vehicle", "Contract Status", "Weekly Payment (₦)",
        "Total Paid (₦)", "Extra Expenditure (₦)", "Outstanding Balance (₦)", "Date Registered",
    ]

    def _flat(r):
        d, c = r["driver"], r["contract"]
        nok  = f"{d.nok_name} ({d.nok_relationship})" if d.nok_name else ""
        return [
            d.full_name, d.phone, d.address or "", nok,
            d.guarantor1_name or "",
            r["vehicle"].vehicle_number if r["vehicle"] else "",
            c.status.title() if c else "—",
            _fmt(c.weekly_amount) if c else "0.00",
            _fmt(r["total_paid"]), _fmt(r["total_exp"]), _fmt(r["outstanding"]),
            d.date_registered.strftime("%Y-%m-%d") if d.date_registered else "",
        ]

    if fmt == "csv":
        return _csv_resp(HDRS, [_flat(r) for r in rows], f"thms_drivers_{_today()}.csv")

    if fmt == "excel":
        return _xlsx_resp("Driver Report", HDRS, [_flat(r) for r in rows],
                          f"thms_drivers_{_today()}.xlsx",
                          totals=[
                              ("Total Drivers",              len(rows)),
                              ("Total Paid (All Drivers)",   f"₦{total_paid:,.2f}"),
                              ("Total Outstanding",          f"₦{total_outstanding:,.2f}"),
                          ])

    if fmt == "pdf":
        return _pdf_resp("reports/pdf/driver.html", dict(
            rows=rows, search=search, status_f=status_f,
            total_paid=total_paid, total_outstanding=total_outstanding,
            now=_now(),
        ), f"thms_drivers_{_today()}.pdf")

    return render_template("reports/driver.html",
        pag=pag, rows=rows, search=search, status_f=status_f,
        total_paid=total_paid, total_outstanding=total_outstanding,
        now=_now(),
    )


# ══════════════════════════════════════════════════════════════════════════════
# 2. VEHICLE REPORT
# ══════════════════════════════════════════════════════════════════════════════

@reports_bp.route("/vehicle")
@login_required
def vehicle():
    search   = request.args.get("q", "").strip()
    status_f = request.args.get("status", "")
    fmt      = request.args.get("format", "")
    page     = request.args.get("page", 1, type=int)

    q = Vehicle.query
    if search:
        like = f"%{search}%"
        q = q.filter(
            Vehicle.vehicle_number.ilike(like) |
            Vehicle.manufacturer.ilike(like) |
            Vehicle.model.ilike(like)
        )
    if status_f:
        q = q.filter(Vehicle.status == status_f)
    q = q.order_by(Vehicle.vehicle_number)

    is_export = bool(fmt)
    pag   = None if is_export else q.paginate(page=page, per_page=50, error_out=False)
    items = q.all() if is_export else pag.items

    def _enrich(v):
        income = float(
            db.session.query(func.coalesce(func.sum(Payment.amount), 0))
            .join(Contract, Payment.contract_id == Contract.id)
            .filter(Contract.vehicle_id == v.id, Payment.is_archived == False)
            .scalar() or 0
        )
        repairs = float(
            db.session.query(func.coalesce(func.sum(Expense.amount), 0))
            .filter(
                Expense.vehicle_id == v.id,
                Expense.is_archived == False,
                Expense.category.in_(["vehicle_repairs", "accident_repairs",
                                       "servicing", "spare_parts"]),
            ).scalar() or 0
        )
        total_exp = float(
            db.session.query(func.coalesce(func.sum(Expense.amount), 0))
            .filter(Expense.vehicle_id == v.id, Expense.is_archived == False)
            .scalar() or 0
        )
        c = v.active_contract
        return dict(vehicle=v, contract=c,
                    driver=c.driver if c else None,
                    income=income, repairs=repairs,
                    total_exp=total_exp,
                    outstanding=c.outstanding_balance if c else 0.0)

    rows              = [_enrich(v) for v in items]
    total_income      = sum(r["income"]      for r in rows)
    total_expenditure = sum(r["total_exp"]   for r in rows)
    total_outstanding = sum(r["outstanding"] for r in rows)

    HDRS = [
        "Plate / Fleet No.", "Make", "Model", "Year", "Purchase Cost (₦)",
        "Current Driver", "Contract Status",
        "Total Income (₦)", "Total Repairs (₦)", "Total Expenditure (₦)", "Outstanding (₦)",
    ]

    def _flat(r):
        v, c = r["vehicle"], r["contract"]
        return [
            v.vehicle_number, v.manufacturer or "", v.model or "",
            v.year or "", _fmt(v.purchase_price),
            r["driver"].full_name if r["driver"] else "—",
            c.status.title() if c else "No Contract",
            _fmt(r["income"]), _fmt(r["repairs"]),
            _fmt(r["total_exp"]), _fmt(r["outstanding"]),
        ]

    if fmt == "csv":
        return _csv_resp(HDRS, [_flat(r) for r in rows], f"thms_vehicles_{_today()}.csv")

    if fmt == "excel":
        return _xlsx_resp("Vehicle Report", HDRS, [_flat(r) for r in rows],
                          f"thms_vehicles_{_today()}.xlsx",
                          totals=[
                              ("Total Vehicles",          len(rows)),
                              ("Total Income Generated",  f"₦{total_income:,.2f}"),
                              ("Total Expenditure",       f"₦{total_expenditure:,.2f}"),
                              ("Total Outstanding",       f"₦{total_outstanding:,.2f}"),
                          ])

    if fmt == "pdf":
        return _pdf_resp("reports/pdf/vehicle.html", dict(
            rows=rows, search=search, status_f=status_f,
            total_income=total_income,
            total_expenditure=total_expenditure,
            total_outstanding=total_outstanding,
            now=_now(),
        ), f"thms_vehicles_{_today()}.pdf")

    return render_template("reports/vehicle.html",
        pag=pag, rows=rows, search=search, status_f=status_f,
        total_income=total_income,
        total_expenditure=total_expenditure,
        total_outstanding=total_outstanding,
        now=_now(),
    )


# ══════════════════════════════════════════════════════════════════════════════
# 3. CONTRACT REPORT
# ══════════════════════════════════════════════════════════════════════════════

@reports_bp.route("/contract")
@login_required
def contract():
    search   = request.args.get("q", "").strip()
    status_f = request.args.get("status", "active")
    fmt      = request.args.get("format", "")
    page     = request.args.get("page", 1, type=int)

    q = Contract.query
    if search:
        like = f"%{search}%"
        q = (q.join(Driver,  Contract.driver_id  == Driver.id)
               .join(Vehicle, Contract.vehicle_id == Vehicle.id)
               .filter(Driver.full_name.ilike(like) | Vehicle.vehicle_number.ilike(like)))
    if status_f:
        q = q.filter(Contract.status == status_f)
    q = q.order_by(Contract.id.desc())

    status_counts = {
        s: Contract.query.filter_by(status=s).count()
        for s in ("active", "suspended", "inactive", "completed", "archived")
    }

    is_export = bool(fmt)
    pag   = None if is_export else q.paginate(page=page, per_page=50, error_out=False)
    items = q.all() if is_export else pag.items

    total_paid        = sum(c.total_paid           for c in items)
    total_outstanding = sum(c.outstanding_balance  for c in items)

    HDRS = [
        "Contract #", "Driver", "Vehicle", "Status",
        "Start Date", "End Date", "Weekly Amount (₦)", "Duration (Yrs)",
        "Weeks Paid", "Weeks Remaining", "Progress %",
        "Total Paid (₦)", "Outstanding (₦)",
    ]

    def _flat(c):
        return [
            f"#{c.id}",
            c.driver.full_name          if c.driver          else "—",
            c.vehicle.vehicle_number    if c.vehicle         else "—",
            c.status.title(),
            c.start_date.strftime("%Y-%m-%d")        if c.start_date        else "—",
            c.expected_end_date.strftime("%Y-%m-%d") if c.expected_end_date else "—",
            _fmt(c.weekly_amount), c.years_agreed or "—",
            c.weeks_completed, c.weeks_remaining,
            f"{c.progress_percent}%",
            _fmt(c.total_paid), _fmt(c.outstanding_balance),
        ]

    if fmt == "csv":
        return _csv_resp(HDRS, [_flat(c) for c in items], f"thms_contracts_{_today()}.csv")

    if fmt == "excel":
        return _xlsx_resp("Contract Report", HDRS, [_flat(c) for c in items],
                          f"thms_contracts_{_today()}.xlsx",
                          totals=[
                              ("Total Contracts",   len(items)),
                              ("Total Paid",        f"₦{total_paid:,.2f}"),
                              ("Total Outstanding", f"₦{total_outstanding:,.2f}"),
                          ])

    if fmt == "pdf":
        return _pdf_resp("reports/pdf/contract.html", dict(
            items=items, search=search, status_f=status_f,
            status_counts=status_counts,
            total_paid=total_paid, total_outstanding=total_outstanding,
            now=_now(),
        ), f"thms_contracts_{_today()}.pdf")

    return render_template("reports/contract.html",
        pag=pag, items=items, search=search, status_f=status_f,
        status_counts=status_counts,
        total_paid=total_paid, total_outstanding=total_outstanding,
        now=_now(),
    )


# ══════════════════════════════════════════════════════════════════════════════
# 4. PAYMENT REPORT
# ══════════════════════════════════════════════════════════════════════════════

@reports_bp.route("/payment")
@login_required
def payment():
    period      = request.args.get("period", "monthly")
    date_from_s = request.args.get("date_from", "")
    date_to_s   = request.args.get("date_to", "")
    method_f    = request.args.get("method", "")
    search      = request.args.get("q", "").strip()
    fmt         = request.args.get("format", "")
    page        = request.args.get("page", 1, type=int)

    date_from, date_to = _period_dates(period, date_from_s, date_to_s)
    df_s = date_from.isoformat() if date_from else ""
    dt_s = date_to.isoformat()   if date_to   else ""

    q = Payment.query.filter(Payment.is_archived == False)
    if date_from:
        q = q.filter(Payment.payment_date >= date_from)
    if date_to:
        q = q.filter(Payment.payment_date <= date_to)
    if method_f:
        q = q.filter(Payment.payment_method == method_f)
    if search:
        like = f"%{search}%"
        q = (q.join(Driver, Payment.driver_id == Driver.id)
               .filter(Driver.full_name.ilike(like) | Payment.receipt_number.ilike(like)))
    q = q.order_by(Payment.payment_date.desc(), Payment.id.desc())

    # Aggregates from the full (untruncated) filtered query
    total_amount = float(q.with_entities(func.sum(Payment.amount)).scalar() or 0)
    total_count  = q.count()

    if date_from and date_to:
        days       = max((date_to - date_from).days + 1, 7)
        avg_weekly = total_amount / (days / 7)
    else:
        avg_weekly = 0.0

    users_map = {u.id: u.username for u in User.query.all()}

    is_export = bool(fmt)
    pag   = None if is_export else q.paginate(page=page, per_page=50, error_out=False)
    items = q.all() if is_export else pag.items

    HDRS = [
        "Receipt #", "Payment Date", "Driver", "Vehicle",
        "Amount (₦)", "Method", "Reference", "Recorded By",
    ]

    def _flat(p):
        vno = ""
        if p.contract and p.contract.vehicle:
            vno = p.contract.vehicle.vehicle_number
        return [
            p.receipt_number or f"#{p.id}",
            p.payment_date.strftime("%Y-%m-%d") if p.payment_date else "—",
            p.driver.full_name if p.driver else "—",
            vno, _fmt(p.amount), p.method_label,
            p.reference or "", users_map.get(p.recorded_by, "—"),
        ]

    if fmt == "csv":
        return _csv_resp(HDRS, [_flat(p) for p in items], f"thms_payments_{_today()}.csv")

    if fmt == "excel":
        return _xlsx_resp("Payment Report", HDRS, [_flat(p) for p in items],
                          f"thms_payments_{_today()}.xlsx",
                          totals=[
                              ("Total Payments",        len(items)),
                              ("Total Collection",      f"₦{total_amount:,.2f}"),
                              ("Avg Weekly Collection", f"₦{avg_weekly:,.2f}"),
                          ])

    if fmt == "pdf":
        return _pdf_resp("reports/pdf/payment.html", dict(
            items=items, search=search, method_f=method_f,
            period=period, df_s=df_s, dt_s=dt_s,
            total_amount=total_amount, avg_weekly=avg_weekly,
            total_count=total_count, users_map=users_map, now=_now(),
        ), f"thms_payments_{_today()}.pdf")

    return render_template("reports/payment.html",
        pag=pag, items=items, search=search, method_f=method_f,
        period=period, df_s=df_s, dt_s=dt_s,
        total_amount=total_amount, avg_weekly=avg_weekly,
        total_count=total_count, users_map=users_map, now=_now(),
    )


# ══════════════════════════════════════════════════════════════════════════════
# 5. EXPENSE / EXPENDITURE REPORT
# ══════════════════════════════════════════════════════════════════════════════

@reports_bp.route("/expense")
@login_required
def expense():
    search      = request.args.get("q", "").strip()
    category_f  = request.args.get("category", "")
    date_from_s = request.args.get("date_from", "")
    date_to_s   = request.args.get("date_to", "")
    fmt         = request.args.get("format", "")
    page        = request.args.get("page", 1, type=int)

    date_from = _pd(date_from_s)
    date_to   = _pd(date_to_s)

    q = Expense.query.filter(Expense.is_archived == False)
    if search:
        like = f"%{search}%"
        q = q.filter(Expense.title.ilike(like) | Expense.description.ilike(like))
    if category_f:
        q = q.filter(Expense.category == category_f)
    if date_from:
        q = q.filter(Expense.expense_date >= date_from)
    if date_to:
        q = q.filter(Expense.expense_date <= date_to)
    q = q.order_by(Expense.expense_date.desc(), Expense.id.desc())

    # Category totals (full filtered set)
    cat_q = db.session.query(Expense.category, func.sum(Expense.amount))\
        .filter(Expense.is_archived == False)
    if search:
        like = f"%{search}%"
        cat_q = cat_q.filter(Expense.title.ilike(like) | Expense.description.ilike(like))
    if category_f:
        cat_q = cat_q.filter(Expense.category == category_f)
    if date_from:
        cat_q = cat_q.filter(Expense.expense_date >= date_from)
    if date_to:
        cat_q = cat_q.filter(Expense.expense_date <= date_to)
    cat_totals  = {cat: float(amt or 0) for cat, amt in cat_q.group_by(Expense.category).all()}
    grand_total = sum(cat_totals.values())

    is_export = bool(fmt)
    pag   = None if is_export else q.paginate(page=page, per_page=50, error_out=False)
    items = q.all() if is_export else pag.items

    HDRS = [
        "Expense #", "Date", "Driver", "Vehicle",
        "Category", "Title", "Amount (₦)", "Status",
    ]

    def _flat(e):
        return [
            e.expense_number or f"#{e.id}",
            e.expense_date.strftime("%Y-%m-%d") if e.expense_date else "—",
            e.driver.full_name          if e.driver  else "—",
            e.vehicle.vehicle_number    if e.vehicle else "—",
            e.category_label, e.display_title,
            _fmt(e.amount), e.status.replace("_", " ").title(),
        ]

    if fmt == "csv":
        return _csv_resp(HDRS, [_flat(e) for e in items], f"thms_expenses_{_today()}.csv")

    if fmt == "excel":
        cat_summary = [
            (CATEGORIES.get(cat, cat), f"₦{amt:,.2f}")
            for cat, amt in sorted(cat_totals.items())
        ] + [("TOTAL", f"₦{grand_total:,.2f}")]
        return _xlsx_resp("Expense Report", HDRS, [_flat(e) for e in items],
                          f"thms_expenses_{_today()}.xlsx", totals=cat_summary)

    if fmt == "pdf":
        return _pdf_resp("reports/pdf/expense.html", dict(
            items=items, search=search, category_f=category_f,
            date_from_s=date_from_s, date_to_s=date_to_s,
            cat_totals=cat_totals, grand_total=grand_total,
            categories=CATEGORIES, now=_now(),
        ), f"thms_expenses_{_today()}.pdf")

    return render_template("reports/expense.html",
        pag=pag, items=items, search=search,
        category_f=category_f,
        date_from_s=date_from_s, date_to_s=date_to_s,
        cat_totals=cat_totals, grand_total=grand_total,
        categories=CATEGORIES, now=_now(),
    )


# ══════════════════════════════════════════════════════════════════════════════
# 6. CAPITAL REPORT
# ══════════════════════════════════════════════════════════════════════════════

@reports_bp.route("/capital")
@login_required
def capital():
    fmt = request.args.get("format", "")

    summary     = _capital_summary()
    vehicles    = Vehicle.query.filter(Vehicle.purchase_price > 0)\
                         .order_by(Vehicle.purchase_date).all()
    adjustments = CapitalAdjustment.query.order_by(CapitalAdjustment.adjustment_date).all()

    if fmt == "csv":
        hdrs = ["Metric", "Amount (₦)"]
        rows = [
            ("Vehicle Purchase Cost",  _fmt(summary["vehicle_cost"])),
            ("Manual Capital Added",   _fmt(summary["manual_added"])),
            ("Manual Withdrawals",     _fmt(summary["manual_withdrawn"])),
            ("Payments Received",      _fmt(summary["payments_received"])),
            ("Extra Expenditure",      _fmt(summary["extra_expenditure"])),
            ("Net Capital Position",   _fmt(summary["net_capital"])),
            ("Outstanding Contracts",  _fmt(summary["outstanding_balance"])),
        ]
        return _csv_resp(hdrs, rows, f"thms_capital_{_today()}.csv")

    if fmt == "excel":
        hdrs = ["Metric", "Amount (₦)"]
        rows = [
            ("Vehicle Purchase Cost",  f"₦{summary['vehicle_cost']:,.2f}"),
            ("Manual Capital Added",   f"₦{summary['manual_added']:,.2f}"),
            ("Manual Withdrawals",     f"₦{summary['manual_withdrawn']:,.2f}"),
            ("Payments Received",      f"₦{summary['payments_received']:,.2f}"),
            ("Extra Expenditure",      f"₦{summary['extra_expenditure']:,.2f}"),
            ("Net Capital Position",   f"₦{summary['net_capital']:,.2f}"),
            ("Outstanding Balance",    f"₦{summary['outstanding_balance']:,.2f}"),
        ]
        return _xlsx_resp("Capital Report", hdrs, rows, f"thms_capital_{_today()}.xlsx")

    if fmt == "pdf":
        return _pdf_resp("reports/pdf/capital_report.html", dict(
            summary=summary, vehicles=vehicles,
            adjustments=adjustments, now=_now(),
        ), f"thms_capital_{_today()}.pdf")

    return render_template("reports/capital_report.html",
        summary=summary, vehicles=vehicles,
        adjustments=adjustments, now=_now(),
    )


# ══════════════════════════════════════════════════════════════════════════════
# 7. AUDIT REPORT
# ══════════════════════════════════════════════════════════════════════════════

@reports_bp.route("/audit")
@login_required
def audit():
    search      = request.args.get("q", "").strip()
    user_f      = request.args.get("user_id", "")
    module_f    = request.args.get("module", "")
    action_f    = request.args.get("action", "")
    date_from_s = request.args.get("date_from", "")
    date_to_s   = request.args.get("date_to", "")
    fmt         = request.args.get("format", "")
    page        = request.args.get("page", 1, type=int)

    date_from = _pd(date_from_s)
    date_to   = _pd(date_to_s)

    q = AuditLog.query
    if search:
        like = f"%{search}%"
        q = q.filter(AuditLog.description.ilike(like) | AuditLog.action.ilike(like))
    if user_f:
        try:
            q = q.filter(AuditLog.user_id == int(user_f))
        except ValueError:
            pass
    if module_f:
        q = q.filter(AuditLog.entity_type == module_f)
    if action_f:
        q = q.filter(AuditLog.action.ilike(f"%{action_f}%"))
    if date_from:
        q = q.filter(AuditLog.created_at >= datetime.combine(date_from, datetime.min.time()))
    if date_to:
        from datetime import time as _dtime
        q = q.filter(AuditLog.created_at <= datetime.combine(date_to, _dtime(23, 59, 59)))
    q = q.order_by(AuditLog.created_at.desc())

    users   = User.query.order_by(User.username).all()
    modules = sorted({r[0] for r in db.session.query(AuditLog.entity_type).distinct() if r[0]})
    actions = sorted({r[0] for r in db.session.query(AuditLog.action).distinct() if r[0]})
    users_map = {u.id: u.username for u in users}

    is_export = bool(fmt)
    pag   = None if is_export else q.paginate(page=page, per_page=50, error_out=False)
    items = q.all() if is_export else pag.items

    HDRS = [
        "Date & Time", "User", "Action", "Module",
        "Record #", "Description", "IP Address",
    ]

    def _flat(log):
        return [
            log.created_at.strftime("%Y-%m-%d %H:%M:%S") if log.created_at else "—",
            users_map.get(log.user_id, "—"), log.action,
            log.entity_type or "System",
            f"#{log.entity_id}" if log.entity_id else "—",
            log.description or "", log.ip_address or "",
        ]

    if fmt == "csv":
        return _csv_resp(HDRS, [_flat(l) for l in items], f"thms_audit_{_today()}.csv")

    if fmt == "excel":
        return _xlsx_resp("Audit Log Report", HDRS, [_flat(l) for l in items],
                          f"thms_audit_{_today()}.xlsx",
                          totals=[("Total Records", len(items))])

    if fmt == "pdf":
        return _pdf_resp("reports/pdf/audit_report.html", dict(
            items=items, search=search,
            user_f=user_f, module_f=module_f, action_f=action_f,
            date_from_s=date_from_s, date_to_s=date_to_s,
            users_map=users_map, now=_now(),
        ), f"thms_audit_{_today()}.pdf")

    return render_template("reports/audit_report.html",
        pag=pag, items=items, search=search,
        user_f=user_f, module_f=module_f, action_f=action_f,
        date_from_s=date_from_s, date_to_s=date_to_s,
        users=users, modules=modules, actions=actions,
        users_map=users_map, now=_now(),
    )


# ══════════════════════════════════════════════════════════════════════════════
# 8. ARCHIVE REPORT
# ══════════════════════════════════════════════════════════════════════════════

@reports_bp.route("/archive")
@login_required
def archive():
    tab    = request.args.get("tab", "drivers")
    search = request.args.get("q", "").strip()
    fmt    = request.args.get("format", "")
    page   = request.args.get("page", 1, type=int)

    driver_count   = Driver.query.filter(Driver.status == "archived").count()
    vehicle_count  = Vehicle.query.filter(Vehicle.status == "archived").count()
    contract_count = Contract.query.filter(Contract.status == "archived").count()

    items = []
    pag   = None

    if tab == "drivers":
        q = Driver.query.filter(Driver.status == "archived")
        if search:
            like = f"%{search}%"
            q = q.filter(Driver.full_name.ilike(like) | Driver.phone.ilike(like))
        q = q.order_by(Driver.date_archived.desc(), Driver.full_name)
        if fmt:
            items = q.all()
        else:
            pag   = q.paginate(page=page, per_page=50, error_out=False)
            items = pag.items

    elif tab == "vehicles":
        q = Vehicle.query.filter(Vehicle.status == "archived")
        if search:
            like = f"%{search}%"
            q = q.filter(Vehicle.vehicle_number.ilike(like) | Vehicle.manufacturer.ilike(like))
        q = q.order_by(Vehicle.date_archived.desc(), Vehicle.vehicle_number)
        if fmt:
            items = q.all()
        else:
            pag   = q.paginate(page=page, per_page=50, error_out=False)
            items = pag.items

    elif tab == "contracts":
        q = Contract.query.filter(Contract.status == "archived")
        if search:
            like = f"%{search}%"
            q = (q.join(Driver, Contract.driver_id == Driver.id)
                   .filter(Driver.full_name.ilike(like)))
        q = q.order_by(Contract.date_archived.desc(), Contract.id.desc())
        if fmt:
            items = q.all()
        else:
            pag   = q.paginate(page=page, per_page=50, error_out=False)
            items = pag.items

    if fmt == "csv":
        if tab == "drivers":
            hdrs = ["Name", "Phone", "Address", "Status", "Date Registered", "Date Archived"]
            rows = [[d.full_name, d.phone, d.address or "", d.status.title(),
                     d.date_registered.strftime("%Y-%m-%d") if d.date_registered else "",
                     d.date_archived.strftime("%Y-%m-%d")   if d.date_archived   else ""]
                    for d in items]
            return _csv_resp(hdrs, rows, f"thms_archived_drivers_{_today()}.csv")
        if tab == "vehicles":
            hdrs = ["Plate No.", "Make", "Model", "Year", "Purchase Cost (₦)", "Status", "Date Archived"]
            rows = [[v.vehicle_number, v.manufacturer or "", v.model or "",
                     v.year or "", _fmt(v.purchase_price), v.status.title(),
                     v.date_archived.strftime("%Y-%m-%d") if v.date_archived else ""]
                    for v in items]
            return _csv_resp(hdrs, rows, f"thms_archived_vehicles_{_today()}.csv")
        if tab == "contracts":
            hdrs = ["Contract #", "Driver", "Vehicle", "Weekly Amt (₦)", "Total Paid (₦)", "Status", "Date Archived"]
            rows = [[f"#{c.id}",
                     c.driver.full_name       if c.driver  else "—",
                     c.vehicle.vehicle_number if c.vehicle else "—",
                     _fmt(c.weekly_amount), _fmt(c.total_paid), c.status.title(),
                     c.date_archived.strftime("%Y-%m-%d") if c.date_archived else ""]
                    for c in items]
            return _csv_resp(hdrs, rows, f"thms_archived_contracts_{_today()}.csv")

    if fmt == "excel":
        if tab == "drivers":
            hdrs = ["Name", "Phone", "Address", "Status", "Date Registered", "Date Archived"]
            rows = [[d.full_name, d.phone, d.address or "", d.status.title(),
                     d.date_registered.strftime("%Y-%m-%d") if d.date_registered else "",
                     d.date_archived.strftime("%Y-%m-%d")   if d.date_archived   else ""]
                    for d in items]
            return _xlsx_resp("Archived Drivers", hdrs, rows,
                              f"thms_archived_drivers_{_today()}.xlsx",
                              totals=[("Total Archived Drivers", len(rows))])
        if tab == "vehicles":
            hdrs = ["Plate No.", "Make", "Model", "Year", "Purchase Cost (₦)", "Status", "Date Archived"]
            rows = [[v.vehicle_number, v.manufacturer or "", v.model or "",
                     v.year or "", _fmt(v.purchase_price), v.status.title(),
                     v.date_archived.strftime("%Y-%m-%d") if v.date_archived else ""]
                    for v in items]
            return _xlsx_resp("Archived Vehicles", hdrs, rows,
                              f"thms_archived_vehicles_{_today()}.xlsx",
                              totals=[("Total Archived Vehicles", len(rows))])
        if tab == "contracts":
            hdrs = ["Contract #", "Driver", "Vehicle", "Weekly Amt (₦)", "Total Paid (₦)", "Status", "Date Archived"]
            rows = [[f"#{c.id}",
                     c.driver.full_name       if c.driver  else "—",
                     c.vehicle.vehicle_number if c.vehicle else "—",
                     _fmt(c.weekly_amount), _fmt(c.total_paid), c.status.title(),
                     c.date_archived.strftime("%Y-%m-%d") if c.date_archived else ""]
                    for c in items]
            return _xlsx_resp("Archived Contracts", hdrs, rows,
                              f"thms_archived_contracts_{_today()}.xlsx",
                              totals=[("Total Archived Contracts", len(rows))])

    if fmt == "pdf":
        return _pdf_resp("reports/pdf/archive.html", dict(
            tab=tab, items=items, search=search,
            driver_count=driver_count,
            vehicle_count=vehicle_count,
            contract_count=contract_count,
            now=_now(),
        ), f"thms_archive_{tab}_{_today()}.pdf")

    return render_template("reports/archive.html",
        tab=tab, pag=pag, items=items, search=search,
        driver_count=driver_count,
        vehicle_count=vehicle_count,
        contract_count=contract_count,
        now=_now(),
    )
